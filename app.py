from flask import Flask, render_template, request, jsonify, send_file
from database import (
    init_db, add_item, record_sale, get_all_items, 
    get_item_by_id, get_sales_history, delete_item, update_item_quantity
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os

app = Flask(__name__)
app.config['JSON_SORT_KEYS'] = False

# Initialize database on startup
init_db()

@app.route('/')
def dashboard():
    """Display the main dashboard with inventory."""
    items = get_all_items()
    total_items = len(items)
    total_quantity = sum(item['quantity'] for item in items)
    total_value = sum(item['quantity'] * item['cost'] for item in items)
    
    return render_template('dashboard.html', 
                         items=items, 
                         total_items=total_items,
                         total_quantity=total_quantity,
                         total_value=total_value)

@app.route('/add_item', methods=['GET', 'POST'])
def add_item_page():
    """Add a new item to inventory."""
    if request.method == 'POST':
        item_name = request.form.get('item_name', '').strip()
        quantity = request.form.get('quantity', '0')
        cost = request.form.get('cost', '0')
        
        try:
            quantity = int(quantity)
            cost = float(cost)
            
            if not item_name:
                return render_template('add_item.html', error="Item name is required!")
            
            if quantity < 0:
                return render_template('add_item.html', error="Quantity cannot be negative!")
            
            if cost < 0:
                return render_template('add_item.html', error="Cost cannot be negative!")
            
            success, message = add_item(item_name, quantity, cost)
            
            if success:
                return render_template('add_item.html', success=message)
            else:
                return render_template('add_item.html', error=message)
        
        except ValueError:
            return render_template('add_item.html', error="Invalid quantity or cost format!")
    
    return render_template('add_item.html')

@app.route('/add_sale', methods=['GET', 'POST'])
def add_sale_page():
    """Record a sale for an item."""
    items = get_all_items()
    
    if request.method == 'POST':
        item_id = request.form.get('item_id', '')
        quantity_sold = request.form.get('quantity_sold', '0')
        
        try:
            item_id = int(item_id)
            quantity_sold = int(quantity_sold)
            
            if quantity_sold <= 0:
                return render_template('add_sale.html', items=items, error="Quantity must be greater than 0!")
            
            success, message = record_sale(item_id, quantity_sold)
            
            if success:
                return render_template('add_sale.html', items=items, success=message)
            else:
                return render_template('add_sale.html', items=items, error=message)
        
        except ValueError:
            return render_template('add_sale.html', items=items, error="Invalid item or quantity!")
    
    return render_template('add_sale.html', items=items)

@app.route('/sales_history')
def sales_history():
    """Display sales history."""
    sales = get_sales_history()
    return render_template('sales_history.html', sales=sales)

@app.route('/export_excel')
def export_excel():
    """Export inventory to Excel file."""
    items = get_all_items()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Add headers
    headers = ['Item ID', 'Item Name', 'Quantity', 'Cost', 'Total Value', 'Last Update']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment
    
    # Add data
    for item in items:
        total_value = item['quantity'] * item['cost']
        ws.append([
            item['item_id'],
            item['item_name'],
            item['quantity'],
            item['cost'],
            total_value,
            item['last_update']
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=len(items) + 1, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            if cell.column in [3, 4, 5]:  # Quantity, Cost, Total Value
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    
    # Add summary section
    summary_row = len(items) + 3
    ws[f'A{summary_row}'] = "SUMMARY"
    ws[f'A{summary_row}'].font = Font(bold=True, size=11)
    
    ws[f'A{summary_row + 1}'] = "Total Items:"
    ws[f'B{summary_row + 1}'] = len(items)
    
    ws[f'A{summary_row + 2}'] = "Total Quantity:"
    ws[f'B{summary_row + 2}'] = sum(item['quantity'] for item in items)
    
    ws[f'A{summary_row + 3}'] = "Total Value:"
    ws[f'B{summary_row + 3}'] = sum(item['quantity'] * item['cost'] for item in items)
    
    # Save to BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='inventory.xlsx'
    )

@app.route('/api/delete_item/<int:item_id>', methods=['DELETE'])
def api_delete_item(item_id):
    """API endpoint to delete an item."""
    success, message = delete_item(item_id)
    return jsonify({'success': success, 'message': message})

@app.route('/api/update_quantity/<int:item_id>', methods=['POST'])
def api_update_quantity(item_id):
    """API endpoint to update item quantity."""
    data = request.get_json()
    new_quantity = data.get('quantity', 0)
    
    try:
        new_quantity = int(new_quantity)
        if new_quantity < 0:
            return jsonify({'success': False, 'message': 'Quantity cannot be negative!'})
        
        success, message = update_item_quantity(item_id, new_quantity)
        return jsonify({'success': success, 'message': message})
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid quantity!'})

if __name__ == '__main__':
    app.run(debug=True, host='localhost', port=5000)
